from dash import html, dcc, Input, Output, State, dash_table, callback, no_update
import base64
import os
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, date
from bodyloop_sdk.client.client import Client, AuthenticatedClient
from bodyloop_sdk.client.api.authentification import login_api_v2_authentification_token_post
from bodyloop_sdk.client.models.body_login_api_v2_authentification_token_post import BodyLoginApiV2AuthentificationTokenPost

from bodyloop_sdk.client.api.probands import (
    get_probands_api_v2_probands_get,
    get_proband_api_v2_probands_proband_id_get,
    update_proband_api_v2_probands_proband_id_patch,
    create_proband_api_v2_probands_post,
)
from bodyloop_sdk.client.models.proband_data import ProbandData
from bodyloop_sdk.client.api.viatars import (
    get_viatar_api_v2_viatars_viatar_id_get,
)
from bodyloop_sdk.client.api.markers_and_measures import (
    get_heights_api_v2_viatars_viatar_id_heights_get
)


layout = dcc.Tab(
    label="Sync Excel",
    value="sync-excel",
    children=[
        html.Div(
            [
                html.H3("Synchronize Probands and Results with BodyLoop"),
                dcc.Upload(
                    id="upload-file",
                    children=html.Button("Load Excel file"),
                    multiple=False,
                ),
                html.Div(id="upload-info", style={"marginTop": "1rem"}),
                html.Button(
                    "Sync with BodyLoop",
                    id="sync-button",
                    n_clicks=0,
                    disabled=True,
                    style={"marginTop": "1rem", "display": "block"},
                ),
                html.Div(id="sync-info", style={"marginTop": "1rem"}),
                html.Button(
                    "Download",
                    id="download-button",
                    n_clicks=0,
                    disabled=True,
                    style={"marginTop": "1rem", "display": "block"},
                ),
            ],
            style={"padding": "1rem 0"},
        )
    ],
)

stores = [
    dcc.Store(id="stored-file-content"),
    dcc.Store(id="stored-file-name"),
    dcc.Download(id="download-file"),
]


def make_results_filename(filename: str) -> str:
    basename, ext = os.path.splitext(filename)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{basename}__results_{timestamp}{ext}"


def split_player_name(full_name: str) -> tuple[str, str]:
    """name_family shall be the last word, all other words shall be inside name_given"""
    parts = str(full_name).split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return "", parts[0]
    return " ".join(parts[:-1]), parts[-1]


def transform_player_birthday(birthday) -> date:
    default_birthday = date(1900, 1, 1)
    
    if birthday is None:
        return default_birthday

    s = str(birthday).strip()

    if s.isdigit():
        # year-only → decide what to do
        return date(int(s), 1, 1)  # or return None instead

    ts = pd.to_datetime(s, errors="coerce")

    if pd.isna(ts):
        return default_birthday

    return ts.date()

@callback(
    Output("stored-file-content", "data"),
    Output("stored-file-name", "data"),
    Output("upload-info", "children"),
    Output("sync-button", "disabled"),
    Output("download-button", "disabled"),
    Input("upload-file", "contents"),
    State("upload-file", "filename"),
    prevent_initial_call=True,
)
def upload(contents, filename):
    if contents is None or filename is None:
        return None, None, "No file uploaded.", True, True

    # contents looks like: "data:<mime-type>;base64,<base64-data>"
    _, content_string = contents.split(",", 1)

    return (
        content_string,
        filename,
        f"Loaded: {filename}",
        False,
        False,
    )


@callback(
    Output("stored-file-content", "data", allow_duplicate=True),
    Output("sync-info", "children"),
    Input("sync-button", "n_clicks"),
    State("stored-file-content", "data"),
    State("base-url-input", "value"),
    State("username-input", "value"),
    State("password-input", "value"),
    prevent_initial_call=True,
)
def sync(n_clicks, content_string, base_url, username, password):
    if not n_clicks or content_string is None:
        return no_update, no_update

    if not base_url or not username or not password:
        return no_update, "Please provide base URL, username, and password."
    
    try:
        client = Client(
            base_url=base_url,
            verify_ssl=False,
            timeout=3.0
        )

        response = login_api_v2_authentification_token_post.sync_detailed(
            client=client,
            body=BodyLoginApiV2AuthentificationTokenPost(
                grant_type="password",
                username=username,
                password=password
            )
        )
    except Exception as e:
        return no_update, f"Could not connect to {base_url}. Please check the URL and that BodyLoop is running. {e}"

    
    if response.status_code == 200:
        api_token = response.parsed.access_token
    else:
        return no_update, "Login failed. Please check your credentials and try again."

    client = AuthenticatedClient(
        base_url=base_url,
        verify_ssl=False,
        token=api_token, 
        timeout=10.0
    )

    file_bytes = base64.b64decode(content_string)
    players = pd.read_excel(BytesIO(file_bytes), header=1)

    probands = pd.DataFrame([proband.to_dict() for proband in get_probands_api_v2_probands_get.sync(client=client)])
    
    # Additional columns
    players["sync_status"] = pd.Series(dtype="string")
    players["proband_id"] = pd.Series(dtype="Int64")
    players["viatar_id"] = pd.Series(dtype="Int64")
    players["viatar_crtime"] = pd.Series(dtype="string")

    # First pass (read-only): Try to find player by PLAYER_ID and key_external
    for player in players.itertuples():
        matching_probands = probands[probands["key_external"] == player.PLAYER_ID]
        if not matching_probands.empty:
            # If the length is greater than 1, we have multiple matches. In this case, we take the first one and print a warning.
            if len(matching_probands) > 1:
                print(f"Warning: Multiple matches found for {player.PLAYER_NAME} with PLAYER_ID {player.PLAYER_ID}. Taking the first match.")
            players.at[player.Index, 'proband_id'] = matching_probands.iloc[0]['proband_id']
            players.at[player.Index, 'sync_status'] = "Found"

    # Second pass (update): For players that still don't have a proband_id, try to match by name and birthdate
    for player in players[players["proband_id"].isna()].itertuples():
        print(f"{player.PLAYER_NAME} {transform_player_birthday(player.PLAYER_BIRTHDAY)} couldn't be found by its PLAYER_ID. Try to find match by name and birthdate")
        
        (name_given, name_family) = split_player_name(player.PLAYER_NAME)
        print(f"name_given: {name_given} name_family: {name_family}")
        
        matching_probands = probands[(probands["name_given"] == name_given) & (probands["name_family"] == name_family) & (probands["date_of_birth"] == transform_player_birthday(player.PLAYER_BIRTHDAY).isoformat())]
        if not matching_probands.empty:
            players.at[player.Index, 'proband_id'] = matching_probands.iloc[0]['proband_id']
            print(f"Found match for {player.PLAYER_NAME} {transform_player_birthday(player.PLAYER_BIRTHDAY)}. Completing with PLAYER_ID {player.PLAYER_ID}")
            modified_proband = update_proband_api_v2_probands_proband_id_patch.sync(
                client=client,
                proband_id=matching_probands.iloc[0]['proband_id'],
                body=ProbandData(
                    key_external=player.PLAYER_ID
                )
            )
            players.at[player.Index, 'proband_id'] = modified_proband.proband_id
            players.at[player.Index, 'sync_status'] = "Updated PLAYER_ID"

    # Third pass (modify): Update meta-data for existing probands
    for player in players.itertuples():
        if pd.isna(player.proband_id):
            continue
        
        proband = probands[probands["proband_id"] == player.proband_id].iloc[0]
        
        if (f"{proband.name_given} {proband.name_family}" != f"{player.PLAYER_NAME}"):
            print(f"Updating name and birthdate for {player.PLAYER_NAME}")
            (name_given, name_family) = split_player_name(player.PLAYER_NAME)
        
            modified_proband = update_proband_api_v2_probands_proband_id_patch.sync(
                client=client,
                proband_id=player.proband_id,
                body=ProbandData(
                    name_given=name_given,
                    name_family=name_family,
                    date_of_birth=transform_player_birthday(player.PLAYER_BIRTHDAY).isoformat(),
                )
            )
            players.at[player.Index, 'sync_status'] = "Updated"

    # Forth pass (add): For players that still don't have a proband_id, create new probands
    for player in players[players["proband_id"].isna()].itertuples():
        print(f"{player.PLAYER_NAME} {transform_player_birthday(player.PLAYER_BIRTHDAY)} couldn't be found by its PLAYER_ID and name. Creating new proband.")

        (name_given, name_family) = split_player_name(player.PLAYER_NAME)
        created_proband = create_proband_api_v2_probands_post.sync(
            client=client,
            body=ProbandData(
                key_external=player.PLAYER_ID,
                name_given=name_given,
                name_family=name_family,
                date_of_birth=transform_player_birthday(player.PLAYER_BIRTHDAY).isoformat(),
            )
        )
        players.at[player.Index, 'proband_id'] = created_proband.proband_id
        players.at[player.Index, 'sync_status'] = "CREATED"

    # Get the most recent scan for each player and print it out
    for idx, player in players.iterrows():
        specific_proband = get_proband_api_v2_probands_proband_id_get.sync(
            client=client,
            proband_id=player.proband_id
        )
        if specific_proband.viatars:
            max_viatar_id = max(viatar.viatar_id for viatar in specific_proband.viatars)
        else:
            max_viatar_id = None
        players.at[idx, 'viatar_id'] = max_viatar_id

    # Get the heights of the most recent scans for each player and print it out

    columns_to_consider = []
    for idx, player in players.iterrows():
        if pd.isna(player.viatar_id):
            continue
        
        viatar = get_viatar_api_v2_viatars_viatar_id_get.sync(
            client=client,
            viatar_id=player.viatar_id
        )
        players.at[idx, 'viatar_crtime'] = viatar.meta.crtime.strftime("%Y-%m-%d %H:%M:%S")
        
        heights = get_heights_api_v2_viatars_viatar_id_heights_get.sync(
            client=client,
            viatar_id=player.viatar_id
        )
        
        height_dict = {height.height_path: height.height for height in heights}
        
        try:
            column = 'HEIGHT_DIFF_SHOULDERS (cm)'
            if column in players.columns:
                players.at[idx, column] = round((height_dict['arm.acromion.R'] - height_dict['arm.acromion.L']) * 100, 1)
                columns_to_consider.append(column)
        except KeyError:
            pass
        
        try:
            column = 'HEIGHT_DIFF_HIP (cm)'
            if column in players.columns:
                players.at[idx, column] = round((height_dict['torso.asis.R'] - height_dict['torso.asis.L']) * 100, 1)
                columns_to_consider.append(column)
        except KeyError:
            pass
        
        try:
            column = 'HEIGHT_DIFF_KNEE (cm)'
            if column in players.columns:
                players.at[idx, column] = round((height_dict['leg.femur.epicondyle.lateral.R'] - height_dict['leg.femur.epicondyle.lateral.L']) * 100, 1)
                columns_to_consider.append(column)
        except KeyError:
            pass
        
        try:
            column = 'HEIGHT_DIFF_ANKLE (cm)'
            if column in players.columns:
                players.at[idx, column] = round((height_dict['leg.lateral.malleolus.R'] - height_dict['leg.lateral.malleolus.L']) * 100, 1)
                columns_to_consider.append(column)
        except KeyError:
            pass

    columns_to_print = ["PLAYER_ID", "PLAYER_NAME", "PLAYER_BIRTHDAY", "proband_id", "sync_status", "viatar_id", "viatar_crtime"]
    columns_to_print.extend(columns_to_consider)
    print(players[columns_to_print])

    # Write back store-file-content
    workbook = load_workbook(filename=BytesIO(file_bytes))
    worksheet = workbook.active
    
    header_row = 2;

    column_by_header={}
    for column in worksheet.iter_cols(min_row=header_row, max_row=header_row):
        for cell in column:
            column_by_header[cell.value] = cell.column

    row_by_player_id={}
    for row in worksheet.iter_rows(min_row=header_row+1):
        player_id_cell = row[column_by_header["PLAYER_ID"]-1]
        row_by_player_id[player_id_cell.value] = player_id_cell.row

    for idx, player in players.iterrows():
        for column in columns_to_consider:
            value = player[column]
            if pd.notna(value):
                worksheet.cell(row=row_by_player_id[player.PLAYER_ID], column=column_by_header[column], value=value)


    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    synced_content = base64.b64encode(output.read()).decode("utf-8")
    info =  dash_table.DataTable(
        data=players.to_dict("records"),
        columns=[{"name": col, "id": col} for col in columns_to_print],
        page_size=15,
    )
    return (synced_content, info)


@callback(
    Output("download-file", "data"),
    Input("download-button", "n_clicks"),
    State("stored-file-content", "data"),
    State("stored-file-name", "data"),
    prevent_initial_call=True,
)
def download(n_clicks, content_string, filename):
    if not n_clicks or not content_string or not filename:
        return None

    file_bytes = base64.b64decode(content_string)
    download_name = make_results_filename(filename)

    return dcc.send_bytes(lambda buffer: buffer.write(file_bytes), download_name)
